import pandas as pd
from pathlib import Path
import re
import unicodedata
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import sys

# =========================
# 1. CAMINHOS E CONFIGURAÇÕES
# =========================
BASE_DIR = Path(__file__).resolve().parent

# Arquivo Master
ARQ_MASTER = BASE_DIR / "INSPECAO-RIR-PPU.xlsx"

# Nomes exatos das abas (Otimizado: apenas o necessário)
ABA_CALM = "CALM EMITIDAS"
ABA_ESTOQUE = "ESTOQUE"

# =========================
# 1.1 MODOS DA FILA DE PRIORIDADE
# =========================
MODOS_PRIORIDADE = {
    "0": {
        "nome": "RODAR COMO ATUALMENTE",
        "colunas": []
    },
    "1": {
        "nome": "PRIORIDADE POR DIÂMETRO",
        "colunas": ["prio_diam"]
    },
    "2": {
        "nome": "PRIORIDADE POR SOP",
        "colunas": ["prio_sop"]
    },
    "3": {
        "nome": "PRIORIDADE POR SEQUÊNCIA DE MONTAGEM",
        "colunas": ["prio_seq_mont"]
    },
    "4": {
        "nome": "PRIORIDADE POR DIÂMETRO + SOP",
        "colunas": ["prio_diam", "prio_sop"]
    },
    "5": {
        "nome": "PRIORIDADE POR DIÂMETRO + SEQUÊNCIA DE MONTAGEM",
        "colunas": ["prio_diam", "prio_seq_mont"]
    },
    "6": {
        "nome": "PRIORIDADE POR SOP + SEQUÊNCIA DE MONTAGEM",
        "colunas": ["prio_sop", "prio_seq_mont"]
    },
    "7": {
        "nome": "PRIORIDADE POR DIÂMETRO + SOP + SEQUÊNCIA DE MONTAGEM",
        "colunas": ["prio_diam", "prio_sop", "prio_seq_mont"]
    }
}

modo_prioridade = "0"

if len(sys.argv) > 1:
    modo_informado = str(sys.argv[1]).strip()

    if modo_informado in MODOS_PRIORIDADE:
        modo_prioridade = modo_informado
    else:
        print(
            f"Aviso: modo de prioridade '{modo_informado}' não reconhecido. "
            "Será usado o modo atual."
        )

config_prioridade = MODOS_PRIORIDADE[modo_prioridade]

print(
    f"Modo de prioridade selecionado: "
    f"{modo_prioridade} - {config_prioridade['nome']}"
)

# =========================
# 2. FUNÇÕES AUXILIARES
# =========================
def normalizar_texto(txt):
    if pd.isna(txt):
        return ""
    txt = str(txt).strip().upper()
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    return txt


def norm_col(c):
    c = str(c).strip().lower()
    c = unicodedata.normalize("NFKD", c).encode("ascii", "ignore").decode("ascii")
    c = re.sub(r"\s+", "_", c)
    c = re.sub(r"[^\w_]", "", c)
    return c


def preparar_cabecalho(df, idx_header):
    df = df.copy()
    df.columns = df.iloc[idx_header].astype(str).tolist()
    df = df.iloc[idx_header + 1:].reset_index(drop=True)
    df.columns = [norm_col(c) for c in df.columns]

    novas_colunas = []

    for col in df.columns:
        if "codigo" in col and "petrobras" in col:
            novas_colunas.append("codigo_petrobras")
        elif col == "codigo_material":
            novas_colunas.append("codigo_petrobras")
        elif col in ["qtd_necessidade_sp", "qtd_necessidade", "quantidade"]:
            novas_colunas.append("qtd")
        elif col in ["qtd_previs", "qtd_previsto"]:
            novas_colunas.append("qtd_estoque_bruto")
        else:
            novas_colunas.append(col)

    df.columns = novas_colunas

    # Remove colunas totalmente vazias, mas preserva as colunas da fila de prioridade
    colunas_fila_prioridade = {
        "prio_diam",
        "prio_sop",
        "prio_seq_mont"
    }

    mascara_colunas = (
        ~df.isna().all(axis=0)
        | pd.Index(df.columns).isin(colunas_fila_prioridade)
    )

    df = df.loc[:, mascara_colunas]

    return df


def encontrar_linha_cabecalho(df, texto_procurado):
    for i, row in df.iterrows():
        valores = [norm_col(x) for x in row.astype(str).tolist()]
        alvo = norm_col(texto_procurado)
        if alvo in valores:
            return i
    return None


def dedup_cols(cols):
    seen = {}
    out = []
    for c in cols:
        if c not in seen:
            seen[c] = 0
            out.append(c)
        else:
            seen[c] += 1
            out.append(f"{c}_{seen[c]}")
    return out

def limpar_colunas_nan(df):
    df = df.copy()
    df = df.dropna(axis=1, how='all')
    cols_ruins = [
        c for c in df.columns
        if str(c).strip().lower() == 'nan'
        or str(c).strip().lower().startswith('nan_')
    ]
    if cols_ruins:
        df = df.drop(columns=cols_ruins, errors='ignore')
    return df

def carregar_spools_controle_liberados(caminho_arquivo):
    try:
        controle = pd.read_excel(caminho_arquivo, sheet_name="CONTROLE_LIBERADOS", header=None)
        idx = encontrar_linha_cabecalho(controle, "ISOMETRICO_SPOOL")

        if idx is None:
            return set()

        controle = preparar_cabecalho(controle, idx)
        controle.columns = dedup_cols(controle.columns)
        controle.columns = [str(c).strip().lower() for c in controle.columns]

        if 'isometrico_spool' not in controle.columns or 'status' not in controle.columns:
            return set()

        controle['isometrico_spool'] = controle['isometrico_spool'].astype(str).str.strip().str.upper()
        controle['status'] = controle['status'].astype(str).str.strip().str.upper()

        return set(
            controle.loc[
                controle['status'] == 'EM FABRICAÇÃO',
                'isometrico_spool'
            ].dropna()
        )

    except Exception:
        return set()

# =========================
# 3. REGRA DE CLASSIFICAÇÃO
# =========================
def classificar_familia(descricao, codigo=""):
    desc = normalizar_texto(descricao)
    cod = normalizar_texto(codigo)
    base = f"{cod} {desc}"

    if "BOCA DE LOBO" in base:
        return "FABRICAÇÃO"

    if "JUNTA ESPIRAL" in base or ("JUNTA" in base and "ESPIRAL" in base):
        return "MONTAGEM"

    if any(t in base for t in [
        "VALV", "VGA", "GAVETA", "GLOBO", "ESFERA", "RETENCAO", "RETENO", "FILTRO",
        "FIGURA OITO", "FIGURA 8", "RAQUETE", "JUNTA ANEL", "JUNTA OVAL"
    ]):
        return "MONTAGEM"

    if "FLANGE CEGO" in base:
        return "MONTAGEM"

    if any(t in base for t in ["C90", "J45", "CURVA 90", "JOELHO 90", "CURVA 45", "JOELHO"]):
        return "FABRICAÇÃO"

    if "TAMPAO" in base or "TAMPÃO" in base:
        return "FABRICAÇÃO"

    if any(t in base for t in [
        "TUBO", "CURVA", "REDUCAO", "REDUÇÃO",
        "FLANGE", "LUVA", "MEIA-LUVA", "NIPLE", "NIPPLE",
        "CAP", "TAMPO", " TE ", "TEE", "NIP RED", "SOCKOLET", "WELDOLET"
    ]):
        return "FABRICAÇÃO"

    return "VERIFICAR"

# =========================
# 4. LEITURA CENTRALIZADA DAS ABAS
# =========================
print(f"Lendo arquivo unificado: {ARQ_MASTER.name}...")

# 4.1 Aba CALM EMITIDAS
calm_raw = pd.read_excel(ARQ_MASTER, sheet_name=ABA_CALM, header=None)
idx_calm = encontrar_linha_cabecalho(calm_raw, "Código Petrobras")
if idx_calm is None:
    idx_calm = encontrar_linha_cabecalho(calm_raw, "CODIGO_MATERIAL")
    
calm = preparar_cabecalho(calm_raw, idx_calm)
calm.columns = dedup_cols(calm.columns)

# =========================
# 4.1.1 COLUNAS DA FILA DE PRIORIDADE
# =========================
COLUNAS_PRIORIDADE = [
    'prio_diam',
    'prio_sop',
    'prio_seq_mont'
]

colunas_prioridade_encontradas = []

for col in COLUNAS_PRIORIDADE:
    if col in calm.columns:
        calm[col] = pd.to_numeric(calm[col], errors='coerce')
        colunas_prioridade_encontradas.append(col)

if colunas_prioridade_encontradas:
    print(
        "Colunas de prioridade identificadas: "
        + ", ".join(colunas_prioridade_encontradas)
    )
else:
    print(
        "Aviso: nenhuma coluna de prioridade foi encontrada. "
        "O processo continuará no modo atual."
    )

# =========================
# 4.1.2 VALIDAÇÃO DO MODO DE PRIORIDADE
# =========================
colunas_modo_escolhido = config_prioridade["colunas"]

fila_prioridade_ativa = False

if modo_prioridade == "0":
    print("Fila de prioridade: desativada (modo atual).")

elif not colunas_modo_escolhido:
    print("Fila de prioridade: desativada (nenhuma coluna definida para o modo).")

else:
    colunas_ausentes = [
        col for col in colunas_modo_escolhido
        if col not in calm.columns
    ]

    if colunas_ausentes:
        print(
            "Aviso: coluna(s) de prioridade não encontrada(s): "
            + ", ".join(colunas_ausentes)
        )
        print("Fila de prioridade: desativada. Será usado o modo atual.")

    else:
        colunas_com_valor = [
            col for col in colunas_modo_escolhido
            if calm[col].notna().any()
        ]

        if not colunas_com_valor:
            print(
                "Nenhuma prioridade preenchida em: "
                + ", ".join(colunas_modo_escolhido)
            )
            print("Fila de prioridade: desativada. Será usado o modo atual.")

        else:
            fila_prioridade_ativa = True
            print(
                "Prioridades preenchidas identificadas em: "
                + ", ".join(colunas_com_valor)
            )
            print("Fila de prioridade: ativada para esta execução.")

# 4.2 Aba ESTOQUE
estoque_raw = pd.read_excel(ARQ_MASTER, sheet_name=ABA_ESTOQUE, header=None)
idx_estoque = encontrar_linha_cabecalho(estoque_raw, "Código Petrobras")
estoque_df = preparar_cabecalho(estoque_raw, idx_estoque)
estoque_df.columns = dedup_cols(estoque_df.columns)

# =========================
# 5. SEPARAÇÃO DOS FABRICÁVEIS
# =========================
calm['familia'] = calm['familia'].astype(str).str.strip().str.upper()

spools_controle_liberados = carregar_spools_controle_liberados(ARQ_MASTER)

if 'isometrico_spool' in calm.columns:
    calm['isometrico_spool'] = calm['isometrico_spool'].astype(str).str.strip().str.upper()
    calm = calm[~calm['isometrico_spool'].isin(spools_controle_liberados)].copy()

fabricaveis = calm[calm['familia'].eq('FABRICAÇÃO')].copy()


# =========================
# 6. CONSOLIDAÇÃO DO ESTOQUE DISPONÍVEL (PELA QTD PREVISTA)
# =========================
estoque_df['codigo_petrobras'] = estoque_df['codigo_petrobras'].astype(str).str.strip().str.upper()

col_qtd_est = 'qtd_prevista' if 'qtd_prevista' in estoque_df.columns else (
    'qtd_estoque_bruto' if 'qtd_estoque_bruto' in estoque_df.columns else estoque_df.columns[0]
)
estoque_df[col_qtd_est] = pd.to_numeric(estoque_df[col_qtd_est], errors='coerce').fillna(0)

if 'dt_prevista' in estoque_df.columns:
    estoque_df['dt_prevista'] = pd.to_datetime(estoque_df['dt_prevista'], errors='coerce')
else:
    estoque_df['dt_prevista'] = pd.NaT

estoque_df['origem'] = estoque_df.get('origem', '').astype(str).str.strip()

qtd_comprometida_por_codigo = {}
if 'isometrico_spool' in calm.columns:
    calm_comprometidos = calm_raw.copy()
    idx_calm_comp = encontrar_linha_cabecalho(calm_comprometidos, "Código Petrobras")
    if idx_calm_comp is None:
        idx_calm_comp = encontrar_linha_cabecalho(calm_comprometidos, "CODIGO_MATERIAL")

try:
    controle_raw = pd.read_excel(ARQ_MASTER, sheet_name="CONTROLE_LIBERADOS", header=None)
    idx_controle = encontrar_linha_cabecalho(controle_raw, "ISOMETRICO_SPOOL")

    if idx_controle is not None:
        controle = preparar_cabecalho(controle_raw, idx_controle)
        controle.columns = dedup_cols(controle.columns)
        controle.columns = [str(c).strip().lower() for c in controle.columns]

        if 'isometrico_spool' in controle.columns and 'status' in controle.columns:
            controle['isometrico_spool'] = controle['isometrico_spool'].astype(str).str.strip().str.upper()
            controle['status'] = controle['status'].astype(str).str.strip().str.upper()

            spools_ctrl = set(
                controle.loc[controle['status'] == 'EM FABRICAÇÃO', 'isometrico_spool'].dropna()
            )

            calm_comp = calm_raw.copy()
            idx_calm_comp = encontrar_linha_cabecalho(calm_comp, "Código Petrobras")
            if idx_calm_comp is None:
                idx_calm_comp = encontrar_linha_cabecalho(calm_comp, "CODIGO_MATERIAL")

            if idx_calm_comp is not None and spools_ctrl:
                calm_comp = preparar_cabecalho(calm_comp, idx_calm_comp)
                calm_comp.columns = dedup_cols(calm_comp.columns)
                calm_comp.columns = [str(c).strip().lower() for c in calm_comp.columns]

                if 'isometrico_spool' in calm_comp.columns and 'codigo_petrobras' in calm_comp.columns:
                    calm_comp['isometrico_spool'] = calm_comp['isometrico_spool'].astype(str).str.strip().str.upper()
                    calm_comp['codigo_petrobras'] = calm_comp['codigo_petrobras'].astype(str).str.strip().str.upper()
                    calm_comp['qtd'] = pd.to_numeric(calm_comp.get('qtd', 0), errors='coerce').fillna(0)

                    df_comp = calm_comp[calm_comp['isometrico_spool'].isin(spools_ctrl)].copy()
                    if not df_comp.empty:
                        soma = df_comp.groupby('codigo_petrobras', as_index=False)['qtd'].sum()
                        for _, r in soma.iterrows():
                            qtd_comprometida_por_codigo[str(r['codigo_petrobras']).strip().upper()] = float(r['qtd'])
except Exception:
    pass

lotes_estoque = []
for _, row_est in estoque_df.iterrows():
    codigo = str(row_est['codigo_petrobras']).strip().upper()
    origem = str(row_est.get('origem', '')).strip()
    comprometido = qtd_comprometida_por_codigo.get(codigo, 0.0)
    saldo = max(0.0, float(row_est[col_qtd_est]) - comprometido)
    dt_prevista = pd.to_datetime(row_est.get('dt_prevista', pd.NaT), errors='coerce')

    if saldo <= 0:
        continue

    origem_upper = origem.upper()
    is_uda = 'UDA' in origem_upper

    # Regra:
    # - UDA pode ser usado com data atual
    # - demais origens só entram se tiverem DT PREVISTA preenchida
    if (not is_uda) and pd.isna(dt_prevista):
        continue

    lotes_estoque.append({
        'codigo': codigo,
        'origem': origem if origem else 'NÃO INFORMADA',
        'saldo': saldo,
        'dt_prevista': pd.Timestamp.now().normalize() if is_uda else dt_prevista
    })

estoque_simulado_total = {}
qtd_inicial_dict = {}

for lote in lotes_estoque:
    estoque_simulado_total[lote['codigo']] = estoque_simulado_total.get(lote['codigo'], 0.0) + lote['saldo']
    qtd_inicial_dict[lote['codigo']] = qtd_inicial_dict.get(lote['codigo'], 0.0) + lote['saldo']


# =========================
# 7. CRUZAMENTO DE SPOOLS (ALOCAÇÃO DINÂMICA)
# =========================
fabricaveis['codigo_fonte'] = fabricaveis['codigo_petrobras'].astype(str).str.strip().str.upper()
fabricaveis['qtd'] = pd.to_numeric(fabricaveis['qtd'], errors='coerce').fillna(0)

# =========================
# 7.0 DEFINIÇÃO DA ORDEM DE PROCESSAMENTO DOS SPOOLS
# =========================
if 'isometrico_spool' in fabricaveis.columns:
    grupo_col = 'isometrico_spool'
elif 'isometrico' in fabricaveis.columns:
    grupo_col = 'isometrico'
else:
    grupo_col = fabricaveis.columns[0]


# Modo atual: preserva exatamente a ordenação original já usada pelo programa
if not fila_prioridade_ativa:
    fabricaveis = fabricaveis.sort_values(
        [grupo_col, 'codigo_petrobras']
    ).copy()

    print(
        "Ordem do match: modo atual "
        f"({grupo_col} + codigo_petrobras)."
    )

else:
    # Uma linha representativa por spool:
    # como as prioridades são iguais em todas as linhas do mesmo spool,
    # a primeira linha contém a prioridade daquele spool.
    fila_spools = (
        fabricaveis
        .drop_duplicates(subset=[grupo_col], keep='first')
        [[grupo_col] + colunas_modo_escolhido]
        .copy()
    )

    # Segurança: converte novamente as prioridades em número.
    # Campo vazio ou inválido vira NaN e vai para o fim da fila.
    for col in colunas_modo_escolhido:
        fila_spools[col] = pd.to_numeric(
            fila_spools[col],
            errors='coerce'
        )

    # Critérios de ordenação:
    # prioridade escolhida primeiro; spool como desempate estável.
    colunas_ordenacao = colunas_modo_escolhido + [grupo_col]

    fila_spools = fila_spools.sort_values(
        by=colunas_ordenacao,
        ascending=True,
        na_position='last',
        kind='stable'
    ).reset_index(drop=True)

    # Posição explícita da fila para preservar a ordem depois do merge
    fila_spools['ordem_fila_spool'] = range(1, len(fila_spools) + 1)

    # Leva a ordem definida para todas as linhas de materiais do spool
    fabricaveis = fabricaveis.merge(
        fila_spools[[grupo_col, 'ordem_fila_spool']],
        on=grupo_col,
        how='left',
        validate='many_to_one'
    )

    # Ordena primeiramente por spool na fila e, depois, por código de material
    fabricaveis = fabricaveis.sort_values(
        by=['ordem_fila_spool', 'codigo_petrobras'],
        ascending=True,
        na_position='last',
        kind='stable'
    ).copy()

    print(
        "Ordem do match: fila aplicada por "
        + " + ".join(colunas_modo_escolhido)
    )

linhas_resultado = []
linhas_pendente_1_mat = []

data_atual_sistema = pd.Timestamp.now().normalize()

for spool, grp in fabricaveis.groupby(grupo_col, sort=False):
    grp = grp.copy()
    spool_chave = str(spool).strip().upper()

    necessidades = (
        grp.groupby('codigo_fonte', as_index=False)['qtd']
        .sum()
        .rename(columns={'qtd': 'qtd_necessaria_spool'})
    )

    spool_fabricavel = True
    materiais_faltantes = []

    # 7.1 Validação prévia: estoque total válido atende este spool?
    for _, row in necessidades.iterrows():
        cod_val = str(row['codigo_fonte']).strip().upper()
        qtd_necessaria = float(row['qtd_necessaria_spool'])
        disponivel = float(estoque_simulado_total.get(cod_val, 0.0))

        if disponivel < qtd_necessaria:
            spool_fabricavel = False
            materiais_faltantes.append({
                'codigo_fonte': cod_val,
                'qtd_necessaria_spool': qtd_necessaria,
                'qtd_fonte_inicial': disponivel,
                'qtd_faltante': qtd_necessaria - disponivel
            })

    if not spool_fabricavel:
        if len(materiais_faltantes) == 1:
            falta = materiais_faltantes[0]

            grp_pendente = grp.copy()
            grp_pendente = grp_pendente.merge(necessidades, on='codigo_fonte', how='left')
            grp_pendente['qtd_fonte_inicial'] = grp_pendente['codigo_fonte'].map(qtd_inicial_dict).fillna(0)
            grp_pendente['spool_fabricavel'] = 'NÃO'
            grp_pendente['DT PREVISTA'] = pd.NaT
            grp_pendente['Origem'] = ''

            mask_faltante = grp_pendente['codigo_fonte'].eq(falta['codigo_fonte'])
            grp_pendente.loc[mask_faltante, 'Origem'] = 'FALTANDO MATERIAL'
            grp_pendente.loc[~mask_faltante, 'Origem'] = 'OK'

            linhas_pendente_1_mat.append(grp_pendente)

        continue

    mapa_origem_item = {}
    mapa_data_item = {}

    # 7.2 Consumo real e sequencial lote por lote
    for _, row in necessidades.iterrows():
        cod = str(row['codigo_fonte']).strip().upper()
        qtd_necessaria = float(row['qtd_necessaria_spool'])
        qtd_restante = qtd_necessaria

        origens_atendidas = []
        data_final_item = None

        for lote in lotes_estoque:
            if lote['codigo'] == cod and lote['saldo'] > 0:
                data_lote = lote['dt_prevista']

                if pd.notna(data_lote):
                    if data_final_item is None or data_lote > data_final_item:
                        data_final_item = data_lote

                if lote['saldo'] >= qtd_restante:
                    origens_atendidas.append(f"{lote['origem']} ({qtd_restante:.2f})")
                    lote['saldo'] -= qtd_restante
                    qtd_restante = 0.0
                else:
                    origens_atendidas.append(f"{lote['origem']} ({lote['saldo']:.2f})")
                    qtd_restante -= lote['saldo']
                    lote['saldo'] = 0.0

                if qtd_restante <= 0:
                    break

        if not origens_atendidas or data_final_item is None:
            spool_fabricavel = False
            break

        mapa_origem_item[cod] = " + ".join(origens_atendidas)
        mapa_data_item[cod] = data_final_item if data_final_item is not None else pd.NaT

    if not spool_fabricavel:
        continue
    
    grp = grp.merge(necessidades, on='codigo_fonte', how='left')
    grp['qtd_fonte_inicial'] = grp['codigo_fonte'].map(qtd_inicial_dict).fillna(0)
    grp['spool_fabricavel'] = 'SIM'
    grp['Origem'] = grp['codigo_fonte'].map(mapa_origem_item).fillna('')
    grp['DT PREVISTA_DINAMICA'] = grp['codigo_fonte'].map(mapa_data_item)

    linhas_resultado.append(grp)

if linhas_resultado:
    podendo_fabricar = pd.concat(linhas_resultado, ignore_index=True)
else:
    podendo_fabricar = fabricaveis.iloc[0:0].copy()
    podendo_fabricar['qtd_necessaria_spool'] = 0
    podendo_fabricar['qtd_fonte_inicial'] = 0
    podendo_fabricar['spool_fabricavel'] = ''
    podendo_fabricar['Origem'] = ''
    podendo_fabricar['DT PREVISTA_DINAMICA'] = pd.NaT

if linhas_pendente_1_mat:
    pendente_1_mat = pd.concat(linhas_pendente_1_mat, ignore_index=True)
else:
    pendente_1_mat = fabricaveis.iloc[0:0].copy()
    pendente_1_mat['qtd_necessaria_spool'] = 0
    pendente_1_mat['qtd_fonte_inicial'] = 0
    pendente_1_mat['spool_fabricavel'] = ''
    pendente_1_mat['Origem'] = ''
    pendente_1_mat['DT PREVISTA'] = pd.NaT

if 'DT PREVISTA_DINAMICA' in podendo_fabricar.columns:
    podendo_fabricar['DT PREVISTA'] = podendo_fabricar['DT PREVISTA_DINAMICA']
    podendo_fabricar = podendo_fabricar.drop(columns=['DT PREVISTA_DINAMICA'])

if 'DT PREVISTA' in podendo_fabricar.columns and 'spool_fabricavel' in podendo_fabricar.columns:
    col = podendo_fabricar.pop('DT PREVISTA')
    idx = podendo_fabricar.columns.get_loc('spool_fabricavel')
    podendo_fabricar.insert(idx, 'DT PREVISTA', col)

colunas_remover = ['obs', 'familia_nova', 'codigo_fonte']
podendo_fabricar = podendo_fabricar.drop(
    columns=[c for c in colunas_remover if c in podendo_fabricar.columns],
    errors='ignore'
)

colunas_nan = [c for c in podendo_fabricar.columns if str(c).startswith('nan')]
if colunas_nan:
    podendo_fabricar = podendo_fabricar.drop(columns=colunas_nan)

pendente_1_mat = pendente_1_mat.drop(
    columns=[c for c in colunas_remover if c in pendente_1_mat.columns],
    errors='ignore'
)

if 'DT PREVISTA' in podendo_fabricar.columns:
    podendo_fabricar['DT PREVISTA'] = pd.to_datetime(podendo_fabricar['DT PREVISTA'], errors='coerce')

if 'DT PREVISTA' in pendente_1_mat.columns:
    pendente_1_mat['DT PREVISTA'] = pd.to_datetime(pendente_1_mat['DT PREVISTA'], errors='coerce')

if 'isometrico_spool' in podendo_fabricar.columns:
    grupo_spool = 'isometrico_spool'
elif 'isometrico' in podendo_fabricar.columns:
    grupo_spool = 'isometrico'
else:
    grupo_spool = None

if grupo_spool:
    liberacao = (
        podendo_fabricar.groupby(grupo_spool, as_index=False)['DT PREVISTA']
        .max()
        .rename(columns={'DT PREVISTA': 'DT LIBERACAO FABRICACAO'})
    )
    podendo_fabricar = podendo_fabricar.merge(liberacao, on=grupo_spool, how='left')

fabricaveis = limpar_colunas_nan(fabricaveis)
podendo_fabricar = limpar_colunas_nan(podendo_fabricar)
pendente_1_mat = limpar_colunas_nan(pendente_1_mat)

# =========================
# 7.3 NOVA ABA - LIBERADO FABRICAR (MESMA REGRA DO PODENDO, USANDO QTD RECEBIDA)
# =========================
estoque_recebido_df = estoque_df.copy()
estoque_recebido_df['codigo_petrobras'] = estoque_recebido_df['codigo_petrobras'].astype(str).str.strip().str.upper()

col_qtd_receb = None
for c in estoque_recebido_df.columns:
    c_norm = str(c).strip().lower()
    if c_norm in ['qtd_recebido', 'qtd_recebida', 'qtd_receb']:
        col_qtd_receb = c
        break

lotes_estoque_recebido = []
estoque_simulado_total_recebido = {}
qtd_inicial_dict_recebido = {}

if col_qtd_receb is not None:
    estoque_recebido_df[col_qtd_receb] = pd.to_numeric(estoque_recebido_df[col_qtd_receb], errors='coerce').fillna(0)
    estoque_recebido_df['origem'] = estoque_recebido_df.get('origem', '').astype(str).str.strip()

    for _, row_est in estoque_recebido_df.iterrows():
        codigo = str(row_est['codigo_petrobras']).strip().upper()
        origem = str(row_est.get('origem', '')).strip()
        saldo = float(row_est[col_qtd_receb])

        if saldo <= 0:
            continue

        lotes_estoque_recebido.append({
            'codigo': codigo,
            'origem': origem if origem else 'NÃO INFORMADA',
            'saldo': saldo,
            'dt_prevista': pd.Timestamp.now().normalize()
        })

    for lote in lotes_estoque_recebido:
        estoque_simulado_total_recebido[lote['codigo']] = estoque_simulado_total_recebido.get(lote['codigo'], 0.0) + lote['saldo']
        qtd_inicial_dict_recebido[lote['codigo']] = qtd_inicial_dict_recebido.get(lote['codigo'], 0.0) + lote['saldo']

linhas_liberado = []

for spool, grp in fabricaveis.groupby(grupo_col, sort=False):
    grp = grp.copy()
    spool_chave = str(spool).strip().upper()

    necessidades = (
        grp.groupby('codigo_fonte', as_index=False)['qtd']
        .sum()
        .rename(columns={'qtd': 'qtd_necessaria_spool'})
    )

    spool_fabricavel = True

    for _, row in necessidades.iterrows():
        cod_val = str(row['codigo_fonte']).strip().upper()
        qtd_necessaria = float(row['qtd_necessaria_spool'])
        disponivel = float(estoque_simulado_total_recebido.get(cod_val, 0.0))

        if disponivel < qtd_necessaria:
            spool_fabricavel = False
            break

    if not spool_fabricavel:
        continue

    mapa_origem_item = {}
    mapa_data_item = {}

    for _, row in necessidades.iterrows():
        cod = str(row['codigo_fonte']).strip().upper()
        qtd_necessaria = float(row['qtd_necessaria_spool'])
        qtd_restante = qtd_necessaria

        origens_atendidas = []
        data_final_item = None

        for lote in lotes_estoque_recebido:
            if lote['codigo'] == cod and lote['saldo'] > 0:
                data_lote = lote['dt_prevista']

                if pd.notna(data_lote):
                    if data_final_item is None or data_lote > data_final_item:
                        data_final_item = data_lote

                if lote['saldo'] >= qtd_restante:
                    origens_atendidas.append(f"{lote['origem']} ({qtd_restante:.2f})")
                    lote['saldo'] -= qtd_restante
                    qtd_restante = 0.0
                else:
                    origens_atendidas.append(f"{lote['origem']} ({lote['saldo']:.2f})")
                    qtd_restante -= lote['saldo']
                    lote['saldo'] = 0.0

                if qtd_restante <= 0:
                    break

        if not origens_atendidas or data_final_item is None:
            spool_fabricavel = False
            break

        mapa_origem_item[cod] = " + ".join(origens_atendidas)
        mapa_data_item[cod] = data_final_item if data_final_item is not None else pd.NaT

    if not spool_fabricavel:
        continue

    grp = grp.merge(necessidades, on='codigo_fonte', how='left')
    grp['qtd_fonte_inicial'] = grp['codigo_fonte'].map(qtd_inicial_dict_recebido).fillna(0)
    grp['spool_fabricavel'] = 'SIM'
    grp['Origem'] = grp['codigo_fonte'].map(mapa_origem_item).fillna('')
    grp['DT PREVISTA'] = grp['codigo_fonte'].map(mapa_data_item)

    linhas_liberado.append(grp)

if linhas_liberado:
    liberado_fabricar = pd.concat(linhas_liberado, ignore_index=True)
else:
    liberado_fabricar = podendo_fabricar.iloc[0:0].copy()

if 'DT PREVISTA' in liberado_fabricar.columns and 'spool_fabricavel' in liberado_fabricar.columns:
    col = liberado_fabricar.pop('DT PREVISTA')
    idx = liberado_fabricar.columns.get_loc('spool_fabricavel')
    liberado_fabricar.insert(idx, 'DT PREVISTA', col)

liberado_fabricar = liberado_fabricar.drop(
    columns=[c for c in colunas_remover if c in liberado_fabricar.columns],
    errors='ignore'
)

liberado_fabricar = limpar_colunas_nan(liberado_fabricar)

# =========================
# 8. AJUSTE FINAL + GRAVAÇÃO DO RELATÓRIO FINAL
# =========================

# 8.1 Converte datas
if 'DT PREVISTA' in podendo_fabricar.columns:
    podendo_fabricar['DT PREVISTA'] = pd.to_datetime(
        podendo_fabricar['DT PREVISTA'],
        errors='coerce'
    )

if 'DT PREVISTA' in pendente_1_mat.columns:
    pendente_1_mat['DT PREVISTA'] = pd.to_datetime(
        pendente_1_mat['DT PREVISTA'],
        errors='coerce'
    )

if 'DT PREVISTA' in liberado_fabricar.columns:
    liberado_fabricar['DT PREVISTA'] = pd.to_datetime(
        liberado_fabricar['DT PREVISTA'],
        errors='coerce'
    )

if 'DT LIBERACAO FABRICACAO' in podendo_fabricar.columns:
    podendo_fabricar['DT LIBERACAO FABRICACAO'] = pd.to_datetime(
        podendo_fabricar['DT LIBERACAO FABRICACAO'],
        errors='coerce'
    )

if 'DT LIBERACAO FABRICACAO' in liberado_fabricar.columns:
    liberado_fabricar['DT LIBERACAO FABRICACAO'] = pd.to_datetime(
        liberado_fabricar['DT LIBERACAO FABRICACAO'],
        errors='coerce'
    )

# 8.2 Grava arquivo Excel
saida = 'consolidado_match.xlsx'
with pd.ExcelWriter(saida, engine='openpyxl') as writer:
    fabricaveis.to_excel(writer, index=False, sheet_name='FABRICAVEIS')
    podendo_fabricar.to_excel(writer, index=False, sheet_name='PODENDO_FABRICAR')
    pendente_1_mat.to_excel(writer, index=False, sheet_name='PENDENTE DE 1 MAT')
    liberado_fabricar.to_excel(writer, index=False, sheet_name='LIBERADO FABRICAR')

    # =========================
    # Formatação - PODENDO_FABRICAR
    # =========================
    if 'PODENDO_FABRICAR' in writer.sheets:
        worksheet = writer.sheets['PODENDO_FABRICAR']

        if 'DT PREVISTA' in podendo_fabricar.columns:
            col_idx = podendo_fabricar.columns.get_loc('DT PREVISTA') + 1
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value is not None:
                    cell.number_format = 'dd/mm/yyyy'

        if 'DT LIBERACAO FABRICACAO' in podendo_fabricar.columns:
            col_idx = podendo_fabricar.columns.get_loc('DT LIBERACAO FABRICACAO') + 1
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value is not None:
                    cell.number_format = 'dd/mm/yyyy'

    # =========================
    # Formatação - PENDENTE DE 1 MAT
    # =========================
    if 'PENDENTE DE 1 MAT' in writer.sheets:
        worksheet = writer.sheets['PENDENTE DE 1 MAT']

        if 'DT PREVISTA' in pendente_1_mat.columns:
            col_idx = pendente_1_mat.columns.get_loc('DT PREVISTA') + 1
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value is not None:
                    cell.number_format = 'dd/mm/yyyy'

    # =========================
    # Formatação - LIBERADO FABRICAR
    # =========================
    if 'LIBERADO FABRICAR' in writer.sheets:
        worksheet = writer.sheets['LIBERADO FABRICAR']

        # -------------------------
        # Formatação de datas
        # -------------------------
        if 'DT PREVISTA' in liberado_fabricar.columns:
            col_idx = liberado_fabricar.columns.get_loc('DT PREVISTA') + 1
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value is not None:
                    cell.number_format = 'dd/mm/yyyy'

        if 'DT LIBERACAO FABRICACAO' in liberado_fabricar.columns:
            col_idx = liberado_fabricar.columns.get_loc('DT LIBERACAO FABRICACAO') + 1
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value is not None:
                    cell.number_format = 'dd/mm/yyyy'

                
        # =========================
        # RESUMO - LIBERADO FABRICAR
        # =========================
        resumo_df = liberado_fabricar.copy()

        col_spool = None
        for c in resumo_df.columns:
            if str(c).strip().lower() == 'isometrico_spool':
                col_spool = c
                break

        col_diam = None
        for c in resumo_df.columns:
            c_norm = str(c).strip().lower()
            if c_norm in ['diametro_1', 'diametro', 'diametro1']:
                col_diam = c
                break

        col_peso = None
        for c in resumo_df.columns:
            c_norm = str(c).strip().lower()
            if c_norm in ['peso_calculado', 'peso_calculado_com_formula', 'peso_formula']:
                col_peso = c
                break

        if col_peso is None:
            for c in resumo_df.columns:
                c_norm = str(c).strip().lower()
                if c_norm in ['peso_total', 'peso_tota', 'peso']:
                    col_peso = c
                    break

        if col_spool is not None and col_diam is not None and col_peso is not None and not resumo_df.empty:
            resumo_df[col_peso] = pd.to_numeric(resumo_df[col_peso], errors='coerce').fillna(0)

            resumo_spool = (
                resumo_df.groupby(col_spool, as_index=False)
                .agg({
                    col_diam: 'first',
                    col_peso: 'sum'
                })
            )

            resumo_spool = resumo_spool.rename(columns={
                col_diam: 'DIAM',
                col_spool: 'SPOOL',
                col_peso: 'PESO'
            })

            linha_inicio = worksheet.max_row + 4

            worksheet.cell(row=linha_inicio, column=2, value='RESUMO')
            worksheet.cell(row=linha_inicio + 1, column=2, value='DIAM')
            worksheet.cell(row=linha_inicio + 1, column=3, value='SPOOL')
            worksheet.cell(row=linha_inicio + 1, column=4, value='PESO')

            for i, (_, row_resumo) in enumerate(resumo_spool.iterrows(), start=linha_inicio + 2):
                worksheet.cell(row=i, column=2, value=row_resumo['DIAM'])
                worksheet.cell(row=i, column=3, value=row_resumo['SPOOL'])
                worksheet.cell(row=i, column=4, value=float(row_resumo['PESO']))

            linha_total_peso = linha_inicio + 2 + len(resumo_spool)
            linha_total_spool = linha_total_peso + 1

            worksheet.cell(row=linha_total_peso, column=3, value='TOTAL PESO')
            worksheet.cell(row=linha_total_peso, column=4, value=float(resumo_spool['PESO'].sum()))

            worksheet.cell(row=linha_total_spool, column=3, value='TOTAL SPOOL')
            worksheet.cell(row=linha_total_spool, column=4, value=int(resumo_spool['SPOOL'].nunique()))

            for i in range(linha_inicio + 2, linha_total_peso + 1):
                worksheet.cell(row=i, column=4).number_format = '#,##0.00'

            
            # =========================
            # FORMATAÇÃO DO RESUMO
            # =========================
            fill_verde = PatternFill(fill_type='solid', fgColor='1CC7A1')
            fonte_normal = Font(name='Calibri', size=11, bold=False, color='000000')
            fonte_bold = Font(name='Calibri', size=11, bold=True, color='000000')

            borda_fina = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            alinhamento_centro = Alignment(horizontal='center', vertical='center')
            alinhamento_direita = Alignment(horizontal='right', vertical='center')
            alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
            worksheet.cell(row=linha_inicio, column=1).font = fonte_normal

            # Cabeçalho verde: DIAM / SPOOL / PESO
            for col in range(1, 4 + 1):
                cell = worksheet.cell(row=linha_inicio + 1, column=col)
                if col in [2, 3, 4]:
                    cell.fill = fill_verde
                cell.font = fonte_normal
                cell.border = borda_fina
                if col == 2:
                    cell.alignment = alinhamento_direita
                elif col == 3:
                    cell.alignment = alinhamento_esquerda
                elif col == 4:
                    cell.alignment = alinhamento_direita

            # Linhas dos dados
            for row_idx in range(linha_inicio + 2, linha_total_peso):
                for col in range(2, 5):
                    cell = worksheet.cell(row=row_idx, column=col)
                    cell.font = fonte_normal
                    cell.border = borda_fina
                    if col == 2:
                        cell.alignment = alinhamento_direita
                    elif col == 3:
                        cell.alignment = alinhamento_esquerda
                    elif col == 4:
                        cell.alignment = alinhamento_direita

            # Linha TOTAL PESO
            for col in range(3, 5):
                cell = worksheet.cell(row=linha_total_peso, column=col)
                cell.fill = fill_verde
                cell.font = fonte_normal
                cell.border = borda_fina
                cell.alignment = alinhamento_direita if col == 4 else alinhamento_esquerda

            # Linha TOTAL SPOOL
            for col in range(3, 5):
                cell = worksheet.cell(row=linha_total_spool, column=col)
                cell.fill = fill_verde
                cell.font = fonte_normal
                cell.border = borda_fina
                cell.alignment = alinhamento_direita if col == 4 else alinhamento_esquerda

            # Formato numérico
            for row_idx in range(linha_inicio + 2, linha_total_peso + 1):
                worksheet.cell(row=row_idx, column=4).number_format = '#,##0.00'

            worksheet.cell(row=linha_total_spool, column=4).number_format = '0'

    # =========================
    # AJUSTE AUTOMÁTICO DE LARGURA EM TODAS AS ABAS
    # =========================
    for worksheet in writer.book.worksheets:
        for col_cells in worksheet.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter

            for cell in col_cells:
                try:
                    if cell.value is None:
                        continue

                    valor = cell.value
                    if hasattr(valor, 'strftime'):
                        texto = valor.strftime('%d/%m/%Y')
                    else:
                        texto = str(valor)

                    if len(texto) > max_len:
                        max_len = len(texto)
                except:
                    pass

            worksheet.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

print('\nProcesso concluído com sucesso!')
print(f'Relatório gerado em: {saida}')